import os
import numpy as np
import os
import openpyxl


def body_estimation_Allimg(filepath):
    imglist = os.listdir(filepath)
    print("%s:%s" % (str(filepath), str(imglist)))
    return imglist


def print_candidate(candidate):
    print(np.shape(candidate))
    point = ["鼻子", "胸口", "左肩", '左肘', '左手腕', '右肩', '右手腕', '右手腕', '左髋(kuan)骨', '左膝盖', '左脚跟',
             '右髋(kuan)骨', '右膝盖', '右脚跟', '左眼', '右眼', '左耳', '右耳']
    # for i in range(np.shape(candidate)[0]):
    #     print("第%d个特征点的坐标:%s"%(int(candidate[i][3]),str((candidate[i][0],candidate[i][1]))))
    for i in range(min(14, np.shape(candidate)[0])):
        print("%s坐标:%s" % (point[i], str((candidate[i][0], candidate[i][1]))))


def writedown(path, candidate):
    os.chdir(path)  # 修改工作路径
    workbook = openpyxl.load_workbook('points.xlsx')  # 返回一个workbook数据类型的值
    sheet = workbook.active  # 获取活动表
    # print('当前活动表是：' + str(sheet))

    for i in range(np.shape(candidate)[0]):
        # sheet.append([candidate[i][0],candidate[i][1]])
        sheet.cell(row=i + 1, column=2 * i + 1).value = candidate[i][0]
        sheet.cell(row=i + 1, column=2 * i + 2).value = candidate[i][1]
    workbook.save('points.xlsx')

"""按列写入数据"""
def writedown_list(list, column_name, filename='lengths.xlsx'):
    column_num = {'身高': 2,'id':1}
    if not column_name in column_num.keys():
        print("Error:key not exist!")
        return

    startrow=3      # 以startrow为数据存放的起始行
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for i in range(len(list)):
        sheet.cell(row=i + startrow, column=column_num[column_name]).value = list[i]  # i+x：单元格下标从1开始
    workbook.save(filename)


def candidateright(candidate, subset):  # wyc认为这样写的candidate是正确的所有起码candidateright
    point = ["鼻子", "胸口", "右肩", '右肘', '右手腕', '左肩', '左肘', '左腕', '右髋(kuan)骨', '右膝盖', '右脚跟',
             '右髋(kuan)骨', '左膝盖', '左脚跟', '右眼', '左眼', '右耳', '左耳']
    # print(np.shape(candidate))
    print(np.shape(subset))
    candidate1 = -1 * np.ones((np.shape(subset)[0], 28))  # 定义一个新的数组
    # print(np.shape(candidate))                                #打印数组规模
    # print(np.shape(candidate1))                               #打印数组规模
    for i in range(np.shape(subset)[0]):  # 循环subset  遍历每一个人体
        for j in range(14):  # 遍历subset某一行的所以点位
            if subset[i][j] != -1:  # 如果该点存在   则输出该点坐标
                print("%s%s坐标:%s" % ("第" + str(i + 1) + "个人", point[j],
                                     str((candidate[int(subset[i][j])][0], candidate[int(subset[i][j])][1]))))
                candidate1[i][2 * j] = candidate[int(subset[i][j])][0]
                candidate1[i][2 * j + 1] = candidate[int(subset[i][j])][1]
            else:  # 如果该点不存在  则返回不存在
                print("%s%s坐标:%s" % ("第" + str(i + 1) + "个人", point[j], "不存在"))
    print(candidate1)
    print(subset)
    return candidate1


if __name__ == '__main__':
    list = [2, 4, 6]
    writedown_list(list, '身高')
