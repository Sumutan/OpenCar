import os
import numpy as np
import openpyxl

"""
以下函数负责实现excel的读写
"""
# path = r"C:\Users\85345\Desktop"  根据自己
def writedownExcel(excelpath, candidate, npicture, photoname):
    os.chdir(excelpath)  # 修改工作路径,注意这条可能会引发后续代码工作路径错误
    workbook = openpyxl.load_workbook('points.xlsx')  # 返回一个workbook数据类型的值
    sheet = workbook.active  # 获取活动表
    # print('当前活动表是：' + str(sheet))

    sheet.cell(row=npicture, column=1).value = photoname
    for i in range(min(14, np.shape(candidate)[0])):
        # sheet.append([candidate[i][0],candidate[i][1]])
        sheet.cell(row=npicture, column=2 * i + 2).value = candidate[i][0]
        sheet.cell(row=npicture, column=2 * i + 3).value = candidate[i][1]
    workbook.save('points.xlsx')


def writedownlength2Excel(excelpath, parameters, npicture):
    start_line = 2
    os.chdir(excelpath)  # 修改工作路径
    workbook = openpyxl.load_workbook('lengths.xlsx')  # 返回一个workbook数据类型的值
    sheet = workbook.active  # 获取活动表
    # print('当前活动表是：' + str(sheet))
    for i in range(len(parameters)):
        sheet.cell(row=npicture + start_line, column=i + 1).value = parameters[i]
    workbook.save('lengths.xlsx')


if __name__ == '__main__':
    excelpath = './'
    npicture = 1

    for i in range(1,10):
        parameters = ["{}.jpg".format(i+1), 213, 432, 133, 123.4]
        writedownlength2Excel(excelpath, parameters, npicture)  # npicture从1开始
        npicture+=1

